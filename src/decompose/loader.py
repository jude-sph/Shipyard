import logging
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def _strip(val) -> str:
    """Convert cell value to stripped string. Handle None, floats, non-breaking spaces."""
    if val is None:
        return ""
    s = str(val).strip()
    s = s.replace("\xa0", " ")
    return s


@dataclass
class WorkbookData:
    digs: dict[str, dict] = field(default_factory=dict)
    system_hierarchy: list[dict] = field(default_factory=list)
    gtr_chapters: list[str] = field(default_factory=list)
    sds_chapters: list[str] = field(default_factory=list)
    acceptance_phases: list[str] = field(default_factory=list)
    verification_methods: list[dict] = field(default_factory=list)
    verification_events: list[dict] = field(default_factory=list)


def load_workbook_data(xlsx_path: Path) -> WorkbookData:
    logger.info(f"Loading workbook: {xlsx_path}")
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    data = WorkbookData()
    sheets = set(wb.sheetnames)

    # Auto-detect format based on sheet names
    if "DIG_Decomp" in sheets:
        logger.info("Detected new format (DIG_Decomp/GTR_Chp/SDS_Chap/PBD)")
        _load_new_format(wb, data, sheets)
    elif "Requirements Decomposition" in sheets:
        logger.info("Detected original format (Requirements Decomposition/System Hierarchy/...)")
        _load_original_format(wb, data, sheets)
    else:
        wb.close()
        raise ValueError(
            "Unrecognised workbook format. Expected either 'DIG_Decomp' sheet "
            "(new format) or 'Requirements Decomposition' sheet (original format)."
        )

    wb.close()
    logger.info(
        f"Loaded: {len(data.digs)} DIGs, {len(data.system_hierarchy)} hierarchy entries, "
        f"{len(data.gtr_chapters)} GTR chapters, {len(data.sds_chapters)} SDS chapters, "
        f"{len(data.acceptance_phases)} phases, {len(data.verification_methods)} methods, "
        f"{len(data.verification_events)} events"
    )
    return data


# =========================================================================
# Original format (7 sheets)
# =========================================================================

def _load_original_format(wb, data: WorkbookData, sheets: set) -> None:
    _load_digs_original(wb["Requirements Decomposition"], data)
    if "System Hierarchy" in sheets:
        _load_system_hierarchy_flat(wb["System Hierarchy"], data)
    if "GTR Chapters" in sheets:
        _load_chapters(wb["GTR Chapters"], data.gtr_chapters, prefix="GTR-Ch")
    if "SDS Chapters" in sheets:
        _load_chapters(wb["SDS Chapters"], data.sds_chapters, prefix="SDS-Ch")
    if "Acceptance Phases" in sheets:
        _load_acceptance_phases(wb["Acceptance Phases"], data)
    if "Verification Events" in sheets:
        _load_verification_methods(wb["Verification Events"], data)
    if "Verification Means" in sheets:
        _load_verification_events(wb["Verification Means"], data)


def _load_digs_original(ws, data: WorkbookData) -> None:
    seen_ids = set()
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] is None:
            continue
        raw_id = _strip(row[0])
        if not raw_id:
            continue
        try:
            dig_id = str(int(float(raw_id)))
        except (ValueError, TypeError):
            logger.warning(f"Row {i}: cannot parse DIG ID '{raw_id}' as a number, skipping")
            continue
        dig_text = _strip(row[1])
        if not dig_text:
            continue
        if dig_id in seen_ids:
            logger.warning(f"Duplicate DIG ID {dig_id} at row {i}, skipping")
            continue
        seen_ids.add(dig_id)
        data.digs[dig_id] = {"dig_id": dig_id, "dig_text": dig_text}
    logger.info(f"  Loaded {len(data.digs)} unique DIGs")


def _load_system_hierarchy_flat(ws, data: WorkbookData) -> None:
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        val = _strip(row[0])
        if val:
            data.system_hierarchy.append({"id": val})


# =========================================================================
# New format (4 sheets: DIG_Decomp, GTR_Chp, SDS_Chap, PBD)
# =========================================================================

def _load_new_format(wb, data: WorkbookData, sheets: set) -> None:
    _load_digs_new(wb["DIG_Decomp"], data)
    if "GTR_Chp" in sheets:
        _load_chapters(wb["GTR_Chp"], data.gtr_chapters, prefix="GTR-Ch")
    if "SDS_Chap" in sheets:
        _load_chapters(wb["SDS_Chap"], data.sds_chapters, prefix="SDS-Ch")
    if "PBD" in sheets:
        _load_pbd_hierarchy(wb["PBD"], data)


def _load_digs_new(ws, data: WorkbookData) -> None:
    """Load DIGs from DIG_Decomp sheet (columns: DNG_ID, Primary_Text, ...)."""
    seen_ids = set()
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or row[0] is None:
            continue
        raw_id = _strip(row[0])
        if not raw_id:
            continue
        try:
            dig_id = str(int(float(raw_id)))
        except (ValueError, TypeError):
            logger.warning(f"Row {i}: cannot parse DIG ID '{raw_id}' as a number, skipping")
            continue
        dig_text = _strip(row[1]) if len(row) > 1 else ""
        if not dig_text:
            continue
        if dig_id in seen_ids:
            logger.warning(f"Duplicate DIG ID {dig_id} at row {i}, skipping")
            continue
        seen_ids.add(dig_id)
        data.digs[dig_id] = {"dig_id": dig_id, "dig_text": dig_text}
    logger.info(f"  Loaded {len(data.digs)} unique DIGs")


def _load_pbd_hierarchy(ws, data: WorkbookData) -> None:
    """Load Product Breakdown hierarchy from PBD sheet.

    The PBD sheet has columns LvL 0 through LvL 5. Each row has a value
    in exactly one column, indicating its level. We flatten this into a
    list of hierarchy entries with their ID strings.
    """
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        # Find the first non-empty cell — that's the hierarchy entry
        for col_idx in range(min(len(row), 6)):
            val = _strip(row[col_idx])
            if val:
                data.system_hierarchy.append({"id": val})
                break
    logger.info(f"  Loaded {len(data.system_hierarchy)} hierarchy entries from PBD")


# =========================================================================
# Shared loaders (used by both formats)
# =========================================================================

def _load_chapters(ws, chapter_list: list[str], prefix: str) -> None:
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        val = _strip(row[0])
        if val and val.startswith(prefix):
            chapter_list.append(val)


def _load_acceptance_phases(ws, data: WorkbookData) -> None:
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row:
            continue
        val = _strip(row[0])
        if val and "Acceptance Phase" in val:
            data.acceptance_phases.append(val)


def _load_verification_methods(ws, data: WorkbookData) -> None:
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        name = _strip(row[0])
        desc = _strip(row[1]) if len(row) > 1 else ""
        if name:
            data.verification_methods.append({"name": name, "description": desc})


def _load_verification_events(ws, data: WorkbookData) -> None:
    current_name = ""
    current_desc = ""
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        name = _strip(row[0])
        desc = _strip(row[1]) if len(row) > 1 else ""
        if name:
            if current_name:
                data.verification_events.append({"name": current_name, "description": current_desc})
            current_name = name
            current_desc = desc
        elif current_name and desc:
            current_desc += " " + desc
    if current_name:
        data.verification_events.append({"name": current_name, "description": current_desc})
